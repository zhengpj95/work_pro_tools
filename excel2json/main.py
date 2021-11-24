import sys
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet
import json
import str2list


def readXlsxFile(xlsxUrl):
    """ 开始读取xlsx文件 """
    wb = load_workbook(filename=xlsxUrl)
    # print(wb.sheetnames)
    # print(wb.worksheets)
    for sheet in wb.worksheets:
        readSingleSheet(sheet)


def getNameList(sheet: worksheet.Worksheet):
    """ 获取当个sheet导出文件的名称以及key数量 """
    serverName = sheet.cell(row=1, column=2).value
    clientName = sheet.cell(row=1, column=5).value
    keyNum = sheet.cell(row=2, column=2).value
    if (not (serverName or clientName)):
        return None
    # print(serverName, clientName, keyNum)
    return {'serverName': serverName, 'clientName': clientName, 'keyNum': keyNum}


def getRowValue(sheet: worksheet.Worksheet, row: int):
    """ 获取某行的数据 """
    columns = sheet.max_column
    rowData = []
    for i in range(1, columns+1):
        cellValue = sheet.cell(row=row, column=i).value
        if (not cellValue):
            break
        rowData.append(cellValue)
    return rowData


def getDataStruct(sheet: worksheet.Worksheet):
    """ 导出数据的结构体信息 """
    row4 = getRowValue(sheet, 4)
    row5 = getRowValue(sheet, 5)
    row6 = getRowValue(sheet, 6)
    row7 = getRowValue(sheet, 7)
    dict = {}
    for i in range(0, len(row4)):
        dict[i] = {
            'name': row5[i],
            'type': row6[i],
            'CS': row7[i]
        }
    # print(dict)
    return dict


def getRealData(sheet: worksheet.Worksheet):
    """ 处理要导出的真实表格数据（多个key的未处理 todo） """
    totalKey = getNameList(sheet)['keyNum']  # key数量
    dataStruct = getDataStruct(sheet)
    maxRow = sheet.max_row
    totalJson = {}
    for i in range(8, maxRow+1):
        rowData = getRowValue(sheet, i)

        # 多个key处理
        eachRowJson = totalJson
        for key in range(0, totalKey):
            if not eachRowJson.get(rowData[key]):
                eachRowJson[rowData[key]] = {}
            eachRowJson = eachRowJson.get(rowData[key])

        for col in range(0, len(rowData)):
            struct = dataStruct[col]
            if 'C' not in struct['CS']:
                continue
            # 特殊处理array, object类型
            if struct['type'] == 'array':
                eachRowJson[struct['name']] = str2list.strToList(rowData[col])
            elif struct['type'] == 'object':
                eachRowJson[struct['name']] = json.loads(rowData[col])
            else:
                eachRowJson[struct['name']] = rowData[col]

    # print(json.dumps(dict, indent=2, ensure_ascii=False))
    dealJsonData(sheet, totalJson)


def dealJsonData(sheet: worksheet.Worksheet, obj: dict):
    """ 导出json数据 """
    nameList = getNameList(sheet)
    if not nameList['clientName']:
        print('xlsx客户端配置名为空')
        return

    with open(nameList['clientName'], "w", encoding='utf-8') as outfile:
        json.dump(obj, outfile, indent=2, ensure_ascii=False)
        # outfile.write(json.dumps(obj, indent=4, ensure_ascii=True))


def readSingleSheet(sheet: worksheet.Worksheet):
    """ 处理单个sheet """
    nameList = getNameList(sheet)
    if (not nameList):
        return
    # print(nameList)
    # print(sheet.max_row, sheet.max_column)
    print('开始处理sheet：', sheet.title)
    getRealData(sheet)


""" start """

if __name__ == '__main__':
    print(sys.argv)
    xlsxUrl = "./test.xlsx"
    if sys.argv and len(sys.argv) > 1 and sys.argv[1]:
        xlsxUrl = sys.argv[1]
    print(xlsxUrl)
    readXlsxFile(xlsxUrl)
