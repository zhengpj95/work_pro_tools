"""
把 excel 导出 json | lua
"""

import sys
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet
import json
import str2list
import time


class SheetStruct:
    """ 表的导出信息 """

    # 服务端导出文件名称
    def serverName(self):
        pass

    # 客户端导出文件名称
    def clientName(self):
        pass

    # key的数量
    def keyCount(self):
        pass

    # 特殊的格式
    def spcialType(self):
        pass


class RowStruct:
    """ 导出数据的结构体信息 """

    # 字段名
    def _name(self):
        pass

    # 类型 (number, string, array, object)
    def _type(self):
        pass

    # 导出字段（S服务端C客户端）
    def _cs(self):
        pass


class Excel2Json:

    xlslUrl = ''
    # 配表信息定义的行数
    structRow = [4, 5, 6, 7]
    # 配表真实数据要导出的列数，每行的后面可能有些说明，但是是不用导出，所以需要记录要导出的真实列数，也就是structRow行的真实列数
    structColLen = 0
    # 配表真实数据开始的行数
    startRow = 8
    sheet: worksheet.Worksheet = None
    # 表头结构体
    sheetStruct: SheetStruct = None

    def __init__(self, xlsxUrl: str) -> None:
        self.xlslUrl = xlsxUrl

    def readFile(self) -> None:
        wb = load_workbook(filename=self.xlslUrl)
        # print(wb.sheetnames)
        # print(wb.worksheets)
        for sheet in wb.worksheets:
            # title中以#开头的表示不导出
            if sheet.title[0] == '#':
                continue
            self.sheet = sheet
            if (not self.getSheetStruct()):
                continue
            self.dealSingleSheet()
        self.sheet = None
        self.xlslUrl = ''
        self.sheetStruct = None

    def getSheetStruct(self) -> SheetStruct:
        """ 获取当个sheet导出文件的名称以及key数量 """
        serverName = self.sheet.cell(row=1, column=2).value
        clientName = self.sheet.cell(row=1, column=5).value
        keyCount = self.sheet.cell(row=2, column=2).value
        specialType = self.sheet.cell(row=2, column=5).value
        if (not (serverName or clientName)):
            return None
        struct = SheetStruct()
        struct.serverName = serverName
        struct.clientName = clientName
        struct.keyCount = keyCount
        struct.spcialType = specialType == 1
        self.sheetStruct = struct
        return struct

    def getRowValue(self, row: int) -> list:
        """ 获取某行的数据 """
        columns = self.sheet.max_column
        rowData = []
        for i in range(1, columns+1):
            cellValue = self.sheet.cell(row=row, column=i).value
            # 单元格填0是需要导出的
            if cellValue == None:
                break
            rowData.append(cellValue)
        return rowData

    def getRowStruct(self) -> dict:
        """ 导出数据的结构体信息 """
        row4 = self.getRowValue(self.structRow[0])
        row5 = self.getRowValue(self.structRow[1])
        row6 = self.getRowValue(self.structRow[2])
        row7 = self.getRowValue(self.structRow[3])
        rowStruct = {}
        self.structColLen = len(row5)
        for i in range(0, len(row5)):
            struct = RowStruct()
            struct._name = row5[i]
            struct._type = row6[i]
            struct._cs = row7[i]
            rowStruct[i] = struct
        return rowStruct

    def dealSingleSheet(self) -> None:
        """ 处理单张sheet """
        if (not self.sheetStruct):
            return
        # print(sheet.max_row, sheet.max_column)
        print('开始处理sheet:', self.sheet.title)
        if self.sheetStruct.spcialType:
            self.dealSpecailReachRowData()
        else:
            self.dealEachRowData()

    def dealSpecailReachRowData(self) -> None:
        """ 处理特殊的导出格式，竖状 """
        rowStruct = self.getRowStruct()
        if not rowStruct:
            return

        totalJson = {}
        for row in range(self.startRow, self.sheet.max_row + 1):
            rowData: RowStruct = self.getRowValue(row)
            if len(rowData) == 0 or rowData[0] == None or 'C' not in rowData[2]:
                break

            totalJson[rowData[0]] = eachRowJson = {}
            col0: RowStruct = rowStruct[0]  # key
            eachRowJson[col0._name] = rowData[0]
            col3: RowStruct = rowStruct[3]  # value

            if rowData[1] == 'array':
                eachRowJson[col3._name] = str2list.strToList(rowData[3])
            elif rowData[1] == 'object':
                eachRowJson[col3._name] = json.loads(rowData[3])
            else:
                eachRowJson[col3._name] = rowData[3]

        self.dealJsonData(totalJson)

    def dealEachRowData(self) -> None:
        """ 读取每行配置，处理导出数据 """
        rowStruct = self.getRowStruct()
        if not rowStruct:
            return

        # 判断是否有客户端字段
        haveClient = False
        for idx in range(0, len(rowStruct)):
            if 'C' in rowStruct[idx]._cs:
                haveClient = True
                break
        if not haveClient:
            print('\t\t不需要导出json')
            return

        totalKey = self.sheetStruct.keyCount  # 表中配置的key数量
        maxRow = self.sheet.max_row

        totalJson = {}
        for row in range(self.startRow, maxRow+1):
            rowData = self.getRowValue(row)
            if len(rowData) == 0 or rowData[0] == None:
                break

            eachRowJson = totalJson
            for key in range(0, totalKey):
                if not eachRowJson.get(rowData[key]):
                    eachRowJson[rowData[key]] = {}
                eachRowJson = eachRowJson.get(rowData[key])

            for col in range(0, self.structColLen):
                colStruct: RowStruct = rowStruct[col]
                if 'C' not in colStruct._cs:
                    continue
                if colStruct._type == 'array':
                    eachRowJson[colStruct._name] = str2list.strToList(
                        rowData[col])
                elif colStruct._type == 'object':
                    eachRowJson[colStruct._name] = json.loads(rowData[col])
                else:
                    eachRowJson[colStruct._name] = rowData[col]
        self.dealJsonData(totalJson)
        # todo
        self.dealLuaData(totalJson)

    def dealJsonData(self, obj: dict) -> None:
        """ 导出json数据 """
        nameList = self.sheetStruct
        if not nameList.clientName:
            print(self.xlslUrl + ' --- 客户端配置名为空 -- 不导出json')
            return

        with open(outputRoot + "/" + nameList.clientName, "w", encoding='utf-8') as outfile:
            json.dump(obj, outfile, indent=2, ensure_ascii=False)
            # outfile.write(json.dumps(obj, indent=4, ensure_ascii=True))

    def dealLuaData(self, obj: dict) -> None:
        """ 导出lua数据 """
        sStruct = self.sheetStruct
        if not sStruct.serverName:
            print(self.xlslUrl + ' --- 服务端配置名为空 -- 不导出lua')
            return
        lua_export_file = open(outputRoot + "/" + sStruct.serverName, 'w')
        lua_export_file.write(
            "-- {0}\n-- {1}\n".format(self.xlslUrl, sStruct.serverName))
        lua_export_file.write(
            "-- %s\n\n" % time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))

        # todo
        luaStr = "return = {\n"
        luaStr += "	[1] = {\n"
        luaStr += "		[1] = {\n"
        luaStr += "			[\"id\"] = 1001\n"
        luaStr += "			[\"name\"] = \"zhangsan\"\n"
        luaStr += "		}\n"
        luaStr += "	}\n"

        lua_export_file.write(luaStr)
        lua_export_file.write("\n}")
        lua_export_file.close()


if __name__ == '__main__':
    print(sys.argv)
    xlsxUrl = "./test.xlsx"
    outputRoot = "./output"
    if sys.argv and len(sys.argv) > 1 and sys.argv[1]:
        xlsxUrl = sys.argv[1]
    if sys.argv and len(sys.argv) > 1 and sys.argv[2]:
        outputRoot = sys.argv[2]

        # print(xlsxUrl)
    excel2Json = Excel2Json(xlsxUrl)
    excel2Json.readFile()
